"""Bytes-level Excel round-trip for the adapy-default procedural engine, plus the
engine-agnostic ``_ADA_META`` sheet writer shared by every engine's exporter.

The adapy-default workbook itself is produced/parsed by :mod:`ada.topo_model.excel`
(via :class:`~ada.topo_model.builder.ProceduralBuilder`). This module wraps that
in the ``bytes -> bytes`` contract the export/import worker jobs want, and stamps
the standard ``_ADA_META`` metadata sheet so an import can auto-detect the engine
that authored the file (see :mod:`ada.comms.rest.procedural`).

Runs on the worker (openpyxl + numpy present), never in the slim API — the API's
side of the round-trip is the dependency-free ``_ADA_META`` *reader* in
``ada.comms.rest.procedural``.
"""

from __future__ import annotations

import datetime as _dt
import pathlib
import tempfile

from ada.comms.rest.procedural import (
    ADA_META_KEY_ENGINE,
    ADA_META_KEY_EXPORTED_AT,
    ADA_META_KEY_META_VERSION,
    ADA_META_KEY_PACKAGE,
    ADA_META_KEY_PACKAGE_VERSION,
    ADA_META_KEY_SCHEMA_VERSION,
    ADA_META_SHEET,
    ADA_META_VERSION,
)
from ada.topo_model.engines import DEFAULT_ENGINE_SLUG, PROCEDURAL_SCHEMA_VERSION

__all__ = [
    "write_ada_meta_sheet",
    "doc_to_xlsx_bytes",
    "xlsx_bytes_to_doc",
]


def write_ada_meta_sheet(
    path: str | pathlib.Path,
    *,
    engine: str,
    package: str,
    package_version: str,
    schema_version: str = PROCEDURAL_SCHEMA_VERSION,
) -> None:
    """Add (or replace) the standard ``_ADA_META`` sheet on an existing xlsx file.

    Vertical key/value layout (A=key, B=value), matching the dependency-free
    reader ``ada.comms.rest.procedural.read_ada_meta_from_xlsx_bytes``. Called by
    every engine's exporter so imports can auto-detect the owning engine."""
    from openpyxl import load_workbook

    wb = load_workbook(path)
    if ADA_META_SHEET in wb.sheetnames:
        del wb[ADA_META_SHEET]
    ws = wb.create_sheet(title=ADA_META_SHEET)
    rows = [
        (ADA_META_KEY_META_VERSION, ADA_META_VERSION),
        (ADA_META_KEY_ENGINE, engine),
        (ADA_META_KEY_PACKAGE, package),
        (ADA_META_KEY_PACKAGE_VERSION, package_version),
        (ADA_META_KEY_SCHEMA_VERSION, schema_version),
        (ADA_META_KEY_EXPORTED_AT, _dt.datetime.now(_dt.timezone.utc).replace(microsecond=0).isoformat()),
    ]
    for r, (key, value) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=value)
    wb.save(path)


def doc_to_xlsx_bytes(doc: dict, *, name: str = "ProceduralModel", engine: str = DEFAULT_ENGINE_SLUG) -> bytes:
    """Serialize a procedural document to the adapy-default multi-sheet workbook,
    with the ``_ADA_META`` sheet stamped. Round-trips :func:`xlsx_bytes_to_doc`."""
    import ada
    from ada.topo_model.builder import ProceduralBuilder

    builder = ProceduralBuilder.from_dict(doc, name=name)
    with tempfile.TemporaryDirectory(prefix="ada-xlsx-export-") as tmp:
        path = pathlib.Path(tmp) / "model.xlsx"
        builder.to_excel(path)
        write_ada_meta_sheet(
            path,
            engine=engine or DEFAULT_ENGINE_SLUG,
            package="ada-py",
            package_version=getattr(ada, "__version__", "unknown"),
            schema_version=builder.schema_version,
        )
        return path.read_bytes()


def xlsx_bytes_to_doc(data: bytes, *, name: str | None = None) -> dict:
    """Parse an adapy-default workbook (bytes) into a procedural document. The
    ``_ADA_META`` sheet is ignored here (it is metadata, not model content) — the
    document's routing header comes from the workbook's ``Model`` sheet."""
    from ada.topo_model.builder import ProceduralBuilder

    with tempfile.TemporaryDirectory(prefix="ada-xlsx-import-") as tmp:
        path = pathlib.Path(tmp) / "model.xlsx"
        path.write_bytes(bytes(data))
        kwargs = {"name": name} if name else {}
        return ProceduralBuilder.from_excel(path, **kwargs).to_doc()
