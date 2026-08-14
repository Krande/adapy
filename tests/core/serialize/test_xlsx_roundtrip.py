"""Extensible pydantic <-> xlsx serializer: round-trip fidelity + extension points."""

from __future__ import annotations

import enum
from typing import Annotated, ClassVar, Literal, Optional

import pytest
from pydantic import BaseModel, Field

from ada.serialize.xlsx import NoCodecError, WorkbookSerializer


class _Color(enum.Enum):
    RED = 1
    BLUE = 2


class Widget(BaseModel):
    SHEET_NAME: ClassVar[str] = "Widgets"
    TAB_COLOR: ClassVar[str] = "92D050"
    HIDE_IN_EXCEL: ClassVar[list[str]] = ["SECRET"]

    NAME: Annotated[str, Field(description="the name")]
    QTY: Annotated[int, Field(description="count")] = 0
    RATIO: Optional[float] = None
    ENABLED: bool = False
    KIND: Literal["a", "b", "c"] = "a"
    ROLE: Literal["only"] = "only"  # single-value constant -> dropped from sheet
    TAGS: Optional[list[str]] = None
    COLOR: Optional[_Color] = None
    SECRET: Optional[str] = None  # hidden, but must still round-trip


def _serializer() -> WorkbookSerializer:
    return WorkbookSerializer().register(Widget)


def test_roundtrip_all_scalar_kinds(tmp_path):
    items = [
        Widget(NAME="w1", QTY=3, RATIO=1.5, ENABLED=True, KIND="b", TAGS=["x", "y,z"], COLOR=_Color.BLUE, SECRET="s"),
        Widget(NAME="w2"),
    ]
    s = _serializer()
    f = tmp_path / "w.xlsx"
    s.write(items, f)
    assert s.read(f)[Widget] == items


def test_comma_in_list_value_survives(tmp_path):
    # The old CSV-join engine corrupted a value containing a comma; the JSON codec
    # keeps it intact.
    items = [Widget(NAME="w", TAGS=["a,b", "c"])]
    s = _serializer()
    f = tmp_path / "w.xlsx"
    s.write(items, f)
    assert s.read(f)[Widget][0].TAGS == ["a,b", "c"]


def test_hidden_field_roundtrips_and_column_is_hidden(tmp_path):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    s = _serializer()
    f = tmp_path / "w.xlsx"
    s.write([Widget(NAME="w", SECRET="keepme")], f)
    assert s.read(f)[Widget][0].SECRET == "keepme"  # not dropped
    ws = load_workbook(f)["Widgets"]
    header = [c.value for c in ws[1]]
    col = header.index("SECRET") + 1
    assert ws.column_dimensions[get_column_letter(col)].hidden is True


def test_single_value_literal_is_not_a_column(tmp_path):
    from openpyxl import load_workbook

    s = _serializer()
    f = tmp_path / "w.xlsx"
    s.write([Widget(NAME="w")], f)
    header = [c.value for c in load_workbook(f)["Widgets"][1]]
    assert "ROLE" not in header  # constant Literal dropped
    assert "KIND" in header  # multi-value Literal kept


def test_multi_value_literal_gets_a_dropdown(tmp_path):
    from openpyxl import load_workbook

    s = _serializer()
    f = tmp_path / "w.xlsx"
    s.write([Widget(NAME="w")], f)
    ws = load_workbook(f)["Widgets"]
    assert any('"a,b,c"' in (dv.formula1 or "") for dv in ws.data_validations.dataValidation)


def test_empty_sheet_written_and_reads_empty(tmp_path):
    s = _serializer()
    f = tmp_path / "w.xlsx"
    s.write([], f)
    from openpyxl import load_workbook

    assert "Widgets" in load_workbook(f).sheetnames
    assert s.read(f)[Widget] == []


def test_unregistered_instance_raises(tmp_path):
    class Other(BaseModel):
        SHEET_NAME: ClassVar[str] = "Others"
        NAME: str

    s = _serializer()
    with pytest.raises(ValueError, match="not registered"):
        s.write([Other(NAME="x")], tmp_path / "w.xlsx")


def test_missing_sheet_name_raises():
    class NoSheet(BaseModel):
        NAME: str

    with pytest.raises(ValueError, match="no SHEET_NAME"):
        WorkbookSerializer().register(NoSheet)


# --- Strict codec policy + custom-codec extension --------------------------
class _Point(BaseModel):
    x: float
    y: float


class Placed(BaseModel):
    SHEET_NAME: ClassVar[str] = "Placed"
    NAME: str
    LOC: _Point  # a nested model — no built-in codec claims it


def test_unclaimed_type_raises_not_silently_dropped(tmp_path):
    s = WorkbookSerializer().register(Placed)
    with pytest.raises(NoCodecError, match="LOC"):
        s.write([Placed(NAME="p", LOC=_Point(x=1, y=2))], tmp_path / "p.xlsx")


def test_custom_codec_by_field_metadata_roundtrips(tmp_path):
    from ada.serialize.xlsx import CodecRegistry, default_registry

    class PointCodec:
        name = "point"

        def matches(self, core):
            return core is _Point

        def to_cell(self, value):
            return None if value is None else f"{value.x},{value.y}"

        def from_cell(self, raw, core):
            if raw is None or raw == "":
                return None
            x, y = (float(v) for v in str(raw).split(","))
            return _Point(x=x, y=y)

    class Placed2(BaseModel):
        SHEET_NAME: ClassVar[str] = "Placed2"
        NAME: str
        LOC: Annotated[_Point, Field(json_schema_extra={"excel": {"codec": "point"}})]

    codecs: CodecRegistry = default_registry().register(PointCodec())
    s = WorkbookSerializer(codecs=codecs).register(Placed2)
    f = tmp_path / "p2.xlsx"
    items = [Placed2(NAME="p", LOC=_Point(x=1.0, y=2.0))]
    s.write(items, f)
    assert s.read(f)[Placed2] == items


def test_jsonlist_codec_handles_list_of_models(tmp_path):
    # A list[NestedModel] via the opt-in jsonlist codec (the system-CONNECTIONS
    # shape): each item dumped to a dict, reconstructed on read.
    class Conn(BaseModel):
        target: str
        kind: Literal["bolt", "weld"]

    class Assembly(BaseModel):
        SHEET_NAME: ClassVar[str] = "Assemblies"
        NAME: str
        CONNECTIONS: Annotated[list[Conn], Field(json_schema_extra={"excel": {"codec": "jsonlist"}})] = []

    s = WorkbookSerializer().register(Assembly)
    f = tmp_path / "a.xlsx"
    items = [Assembly(NAME="a", CONNECTIONS=[Conn(target="t1", kind="bolt"), Conn(target="t2", kind="weld")])]
    s.write(items, f)
    assert s.read(f)[Assembly] == items


# --- Per-model hooks -------------------------------------------------------
def test_should_skip_row_hook(tmp_path):
    class Rowy(BaseModel):
        SHEET_NAME: ClassVar[str] = "Rowy"
        NAME: Optional[str] = None
        VAL: int = 0

        @classmethod
        def excel_should_skip_row(cls, raw: dict) -> bool:
            return not raw.get("NAME")  # drop rows with no NAME

    s = WorkbookSerializer().register(Rowy)
    f = tmp_path / "r.xlsx"
    # Write two rows, one with an empty NAME; the skip hook drops it on read.
    s.write([Rowy(NAME="keep", VAL=1), Rowy(NAME=None, VAL=2)], f)
    read = s.read(f)[Rowy]
    assert [r.NAME for r in read] == ["keep"]


# --- Vertical layout -------------------------------------------------------
def test_vertical_layout_roundtrips(tmp_path):
    class Meta(BaseModel):
        SHEET_NAME: ClassVar[str] = "Meta"
        ORIENTATION: ClassVar[str] = "VERTICAL"
        TITLE: str
        VERSION: int = 1

    s = WorkbookSerializer().register(Meta)
    f = tmp_path / "m.xlsx"
    s.write([Meta(TITLE="demo", VERSION=3)], f)
    assert s.read(f)[Meta] == [Meta(TITLE="demo", VERSION=3)]


# --- Real topology entities ------------------------------------------------
def test_real_topo_entities_roundtrip(tmp_path):
    from ada.topology.entities import EquipRepr, TopoEquipment, TopoOpening, TopoSpace

    spaces = [
        TopoSpace(NAME="Cell1", X=0, Y=0, Z=0, DX=5, DY=5, DZ=3, INCLUDE=True, SE=[0, 1], DESCRIPTION="d"),
        TopoSpace(NAME="Cell2", X=5, Y=0, Z=0, DX=5, DY=5, DZ=3),
    ]
    eqs = [
        TopoEquipment(
            NAME="Pump2",
            DESCRIPTION="pump",
            SPACE_NAME="Cell1",
            SPACE_LOC="FLOOR",
            X=2,
            Y=2,
            Z=0,
            LX=1,
            LY=1,
            LZ=1,
            ROT_Z=-90,
            COGx=0,
            COGy=0,
            COGz=0.5,
            massDry=1000,
            massCont=200,
            CONDITION_NAMES=["a", "b,c"],
            EQ_REPR=EquipRepr.FOOTPRINT_MASS,
        )
    ]
    s = WorkbookSerializer().register(TopoSpace).register(TopoEquipment).register(TopoOpening)
    f = tmp_path / "topo.xlsx"
    s.write([*spaces, *eqs], f)
    back = s.read(f)
    assert back[TopoSpace] == spaces  # incl. hidden DESCRIPTION + SE
    assert back[TopoEquipment] == eqs  # incl. ROT_Z, enum, comma-in-list
    assert back[TopoOpening] == []


def test_alt_sheet_name_alias_reads(tmp_path):
    """A model's ALT_SHEET_NAMES lets a workbook whose sheet is named by a sibling
    tool (TopoSpace's ``Rooms`` alias vs the primary ``Spaces``) still read."""
    import openpyxl

    from ada.topology.entities import TopoSpace

    s = WorkbookSerializer().register(TopoSpace)
    f = tmp_path / "rooms.xlsx"
    s.write([TopoSpace(NAME="C1", X=0, Y=0, Z=0, DX=5, DY=5, DZ=3)], f)
    # Rename the written "Spaces" sheet to the alias "Rooms".
    wb = openpyxl.load_workbook(f)
    wb["Spaces"].title = "Rooms"
    wb.save(f)

    back = s.read(f)
    assert [x.NAME for x in back[TopoSpace]] == ["C1"]
