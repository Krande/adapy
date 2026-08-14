"""Quantity take-off for a compiled :class:`ada.Part`/:class:`ada.Assembly`.

Turns the *structured* model (beams, plates, pipes, routed duct/cable-tray runs)
into a discipline-organised statistics document — the data behind the viewer's
"Stats" panel. The GLB carries only triangles, so these quantities (beam lengths
by section, plate areas by thickness, pipe/duct/tray segment lengths, per-
discipline mass + centre-of-gravity) can only come from the model itself, at
compile time.

Four disciplines are recognised, matching the viewer's four hues:

* ``structural`` — :class:`ada.Beam` (girders/columns/stringers) + :class:`ada.Plate`,
* ``piping``     — :class:`ada.Pipe` (round tube runs, ``IfcPipeSegment``),
* ``hvac``       — box duct runs (``IfcDuctSegment``),
* ``electrical`` — open cable-tray runs (``IfcCableSegment``).

Discipline is read from the routed geometry's ``segment_ifc_class`` metadata (the
tag that survives tessellation) with an object-type fallback, mirroring the
frontend ``classifyMedium``/``classifyEquipment`` heuristics. Everything else
(equipment placeholders, misc shapes) is counted in the object total but is not a
discipline and does not contribute to the discipline mass roll-up.

The output is a plain JSON-able ``dict`` (see :func:`model_takeoff`); the compile
worker stores it gzip-at-rest as a ``.stats.json`` sibling of the procedural GLB.
"""

from __future__ import annotations

from typing import TYPE_CHECKING, Literal

import numpy as np

if TYPE_CHECKING:
    from ada import Part

# Discipline keys — kept in sync with the frontend's four hues.
Discipline = Literal["structural", "piping", "hvac", "electrical", "other"]

DISCIPLINE_NAMES: dict[str, str] = {
    "structural": "Structural",
    "piping": "Piping",
    "hvac": "HVAC",
    "electrical": "Electrical",
}
# The four disciplines shown, in card/legend order.
DISCIPLINE_ORDER: tuple[str, ...] = ("structural", "piping", "hvac", "electrical")

# segment_ifc_class -> discipline (the routed-geometry tag; survives tessellation).
_SEG_CLASS_DISCIPLINE: dict[str, str] = {
    "IfcPipeSegment": "piping",
    "IfcDuctSegment": "hvac",
    "IfcCableSegment": "electrical",
}

_KG_PER_TONNE = 1000.0
SCHEMA_VERSION = 2

# Fabrication-detail joint spec_name -> (slug, display name). The slug matches the
# Detailing engine's advertised joint-type slug (see detailing_catalog.py) so the
# viewer's per-type counts line up; ``adapy.<slug>`` is the connection spec_name.
_JOINT_DISPLAY: dict[str, str] = {
    "girder_gusset": "Girder gusset",
    "column_base_plate": "Column base plate",
    "box_to_box": "Box-to-box",
}


def _joint_slug(spec_name: str | None) -> str:
    """``"adapy.girder_gusset"`` -> ``"girder_gusset"`` (the detailing joint slug)."""
    if not spec_name:
        return "unknown"
    return spec_name.split(".", 1)[1] if "." in spec_name else spec_name


def _joints_takeoff(part: "Part") -> dict:
    """Per-joint overview from the compiled model's fabrication-detail connections.

    Walks every :class:`~ada.api.connections.joints.Connection` parked under the
    detailing stage's ``Part("Joints")`` and reads its ``connection_info`` record
    (name, spec_name, member roles, plate/weld names, centre). Returns a total
    count, a per-type roll-up (``by_type`` — slug/name/count, the data behind the
    Detailing tab's "N detected" badge) and a per-instance ``items`` table (the
    Joints overview). Empty when the model carries no detailing joints.
    """
    from ada.api.connections.joints import Connection

    items: list[dict] = []
    by_type: dict[str, dict] = {}
    for conn in part.get_all_parts_in_assembly(by_type=Connection):
        info = (getattr(conn, "metadata", None) or {}).get("connection_info") or {}
        spec_name = info.get("spec_name") or getattr(conn, "spec_name", None)
        slug = _joint_slug(spec_name)
        name = _JOINT_DISPLAY.get(slug, slug.replace("_", " ").title())
        roles = info.get("member_roles") or {}
        members = sorted({m for names in roles.values() for m in (names or [])})
        plate_names = list(info.get("plate_names") or [])
        weld_names = list(info.get("weld_names") or [])
        row = by_type.setdefault(slug, {"slug": slug, "name": name, "count": 0})
        row["count"] += 1
        items.append(
            {
                "name": info.get("name") or conn.name,
                "slug": slug,
                "type": name,
                "members": members,
                "plates": len(plate_names),
                "welds": len(weld_names),
                "centre": info.get("centre"),
            }
        )

    by_type_rows = sorted(by_type.values(), key=lambda r: r["count"], reverse=True)
    items.sort(key=lambda r: (r["slug"], r["name"]))
    return {"count": len(items), "by_type": by_type_rows, "items": items}


def classify_discipline(obj) -> Discipline:
    """Classify a physical object into one of the four disciplines (or ``other``).

    The routed-run ``segment_ifc_class`` metadata is authoritative (it is the
    discipline tag the router stamps on every duct/tray/pipe run); the object type
    is the fallback for structural steel and bare pipe segments.
    """
    from ada import Beam, Pipe, Plate
    from ada.api.piping.base_piping import PipeSegElbow, PipeSegStraight

    seg_class = (getattr(obj, "metadata", None) or {}).get("segment_ifc_class")
    if seg_class in _SEG_CLASS_DISCIPLINE:
        return _SEG_CLASS_DISCIPLINE[seg_class]  # type: ignore[return-value]

    if isinstance(obj, (Pipe, PipeSegStraight, PipeSegElbow)):
        return "piping"
    if isinstance(obj, Plate):
        return "structural"
    if isinstance(obj, Beam):  # exact Beam and any straight/curved structural beam
        return "structural"
    return "other"


def _rho(obj) -> float:
    """Material density (kg/m^3) for an object, defaulting to steel if absent."""
    mat = getattr(obj, "material", None)
    model = getattr(mat, "model", None)
    rho = getattr(model, "rho", None)
    return float(rho) if rho else 7850.0


def _section_area(section) -> float:
    """Cross-section steel area ``Ax`` (m^2); ``0.0`` if it cannot be resolved."""
    try:
        return float(section.properties.Ax)
    except Exception:
        return 0.0


def _pipe_size_label(section) -> str:
    """Nominal outer-diameter label for a round pipe section (e.g. ``"Ø60 mm"``)."""
    r = getattr(section, "r", None)
    if r:
        return f"Ø{round(float(r) * 2 * 1000):g} mm"
    return section.name


def _box_size_label(section) -> str:
    """``width×height`` label (mm) for a rectangular duct section."""
    w = float(getattr(section, "w_top", 0.0) or 0.0) * 1000
    h = float(getattr(section, "h", 0.0) or 0.0) * 1000
    return f"{w:g}×{h:g}"


def _tray_size_label(section) -> str:
    """Tray-width label (mm). The channel web height carries the tray width."""
    h = float(getattr(section, "h", 0.0) or 0.0) * 1000
    return f"{h:g} mm"


class _Acc:
    """Mass/COG accumulator (kg, mass-weighted position sum)."""

    __slots__ = ("mass", "moment", "count")

    def __init__(self) -> None:
        self.mass = 0.0
        self.moment = np.zeros(3)
        self.count = 0

    def add(self, cog, mass_kg: float) -> None:
        self.mass += mass_kg
        self.moment += np.asarray(cog, dtype=float) * mass_kg
        self.count += 1

    def cog(self) -> list[float]:
        if self.mass <= 1e-12:
            return [0.0, 0.0, 0.0]
        return [round(float(v), 4) for v in (self.moment / self.mass)]

    def mass_t(self) -> float:
        return round(float(self.mass) / _KG_PER_TONNE, 4)


def model_takeoff(part: "Part", *, source_name: str | None = None) -> dict:
    """Compute the discipline-organised quantity take-off for ``part``.

    ``part`` is a compiled :class:`ada.Part`/:class:`ada.Assembly` (the procedural
    builder's output). Returns a plain JSON-able ``dict`` with total mass + COG +
    object count, a per-discipline mass/COG roll-up, and per-discipline take-off
    tables (beams-by-section, plates-by-thickness, pipe/duct/tray segments by size,
    fittings, major items). All lengths are metres, areas m^2, masses tonnes.
    """
    import ada

    disc_acc: dict[str, _Acc] = {k: _Acc() for k in DISCIPLINE_ORDER}
    other_acc = _Acc()

    bbox_min = np.full(3, np.inf)
    bbox_max = np.full(3, -np.inf)

    def _grow_bbox(*points) -> None:
        nonlocal bbox_min, bbox_max
        for p in points:
            arr = np.asarray(p, dtype=float)
            bbox_min = np.minimum(bbox_min, arr)
            bbox_max = np.maximum(bbox_max, arr)

    # Per-object mass for the "major items" table (name, discipline, mass_t, cog).
    major: list[tuple[str, str, float, list[float]]] = []

    # --- structural: beams by section -------------------------------------
    beams_by_section: dict[str, dict] = {}
    for bm in part.get_all_physical_objects(by_type=ada.Beam):
        try:
            cog, mass_kg = bm.get_cog_and_mass()
        except Exception:
            continue
        length = float(bm.length)
        sec = bm.section.name
        row = beams_by_section.setdefault(sec, {"section": sec, "count": 0, "length": 0.0, "mass": 0.0})
        row["count"] += 1
        row["length"] += length
        row["mass"] += mass_kg / _KG_PER_TONNE
        disc_acc["structural"].add(cog, mass_kg)
        _grow_bbox(bm.n1.p, bm.n2.p)
        major.append((bm.name, "structural", mass_kg / _KG_PER_TONNE, [round(float(v), 3) for v in cog]))

    # --- structural: plates by thickness ----------------------------------
    plates_by_thk: dict[float, dict] = {}
    for pl in part.get_all_physical_objects(by_type=ada.Plate):
        try:
            area = float(pl.poly.get_area())
            mass_kg = float(pl.get_mass())
            cog = pl.get_cog()
        except Exception:
            continue
        thk = round(float(pl.t), 6)
        label = f"PL{round(thk * 1000):g}"
        row = plates_by_thk.setdefault(thk, {"label": label, "thickness": thk, "count": 0, "area": 0.0, "mass": 0.0})
        row["count"] += 1
        row["area"] += area
        row["mass"] += mass_kg / _KG_PER_TONNE
        disc_acc["structural"].add(cog, mass_kg)
        for pt in pl.poly.points3d:
            _grow_bbox(pt)
        major.append((pl.name, "structural", mass_kg / _KG_PER_TONNE, [round(float(v), 3) for v in cog]))

    # --- piping: pipes -> segments by nominal size ------------------------
    pipe_by_size: dict[str, dict] = {}
    n_elbows = 0
    n_flanges = 0
    from ada.api.piping.base_piping import PipeSegElbow, PipeSegStraight

    for pipe in part.get_all_physical_objects(by_type=ada.Pipe):
        pipe_mass_kg = 0.0
        pipe_moment = np.zeros(3)
        for seg in pipe.segments:
            rho = _rho(seg)
            area = _section_area(seg.section)
            if isinstance(seg, PipeSegStraight):
                length = float(seg.length)
                mid = (np.asarray(seg.p1.p) + np.asarray(seg.p2.p)) / 2.0
                mass_kg = area * length * rho
                size = _pipe_size_label(seg.section)
                row = pipe_by_size.setdefault(size, {"size": size, "segments": 0, "length": 0.0, "mass": 0.0})
                row["segments"] += 1
                row["length"] += length
                row["mass"] += mass_kg / _KG_PER_TONNE
                disc_acc["piping"].add(mid, mass_kg)
                _grow_bbox(seg.p1.p, seg.p2.p)
            elif isinstance(seg, PipeSegElbow):
                n_elbows += 1
                p1 = np.asarray(seg.p1.p if hasattr(seg.p1, "p") else seg.p1)
                p3 = np.asarray(seg.p3.p if hasattr(seg.p3, "p") else seg.p3)
                length = float(np.linalg.norm(p3 - p1))
                mid = (p1 + p3) / 2.0
                mass_kg = area * length * rho
                disc_acc["piping"].add(mid, mass_kg)
            else:
                continue
            pipe_mass_kg += mass_kg
            pipe_moment += mid * mass_kg
        # flanged end connections ~ endpoints of each pipe run
        n_flanges += 2
        pipe_cog = (pipe_moment / pipe_mass_kg) if pipe_mass_kg > 1e-12 else np.zeros(3)
        major.append((pipe.name, "piping", pipe_mass_kg / _KG_PER_TONNE, [round(float(v), 3) for v in pipe_cog]))

    # --- hvac (duct) + electrical (cable tray): swept runs ----------------
    duct_by_size: dict[str, dict] = {}
    tray_by_size: dict[str, dict] = {}
    for obj in part.get_all_physical_objects():
        seg_class = (getattr(obj, "metadata", None) or {}).get("segment_ifc_class")
        if seg_class not in ("IfcDuctSegment", "IfcCableSegment"):
            continue
        n1 = getattr(obj, "n1", None)
        n2 = getattr(obj, "n2", None)
        if n1 is None or n2 is None:
            continue
        p1 = np.asarray(n1.p)
        p2 = np.asarray(n2.p)
        length = float(np.linalg.norm(p2 - p1))
        mid = (p1 + p2) / 2.0
        section = obj.section
        area_x = _section_area(section)
        rho = _rho(obj)
        mass_kg = area_x * length * rho
        _grow_bbox(p1, p2)
        if seg_class == "IfcDuctSegment":
            size = _box_size_label(section)
            w = float(getattr(section, "w_top", 0.0) or 0.0)
            h = float(getattr(section, "h", 0.0) or 0.0)
            surf = 2.0 * (w + h) * length
            row = duct_by_size.setdefault(size, {"size": size, "segments": 0, "length": 0.0, "area": 0.0, "mass": 0.0})
            row["segments"] += 1
            row["length"] += length
            row["area"] += surf
            row["mass"] += mass_kg / _KG_PER_TONNE
            disc_acc["hvac"].add(mid, mass_kg)
        else:  # IfcCableSegment
            size = _tray_size_label(section)
            row = tray_by_size.setdefault(size, {"size": size, "segments": 0, "length": 0.0, "mass": 0.0})
            row["segments"] += 1
            row["length"] += length
            row["mass"] += mass_kg / _KG_PER_TONNE
            disc_acc["electrical"].add(mid, mass_kg)

    # --- non-discipline objects (equipment, misc shapes): count only ------
    from ada import Shape

    n_objects = 0
    for obj in part.get_all_physical_objects():
        n_objects += 1
        if classify_discipline(obj) == "other":
            mass = getattr(obj, "mass", None)
            if isinstance(obj, Shape) and mass:
                cog = getattr(obj, "cog_abs", None)
                if cog is not None:
                    other_acc.add(cog, float(mass))

    # --- roll-up ----------------------------------------------------------
    def _round_rows(rows: list[dict], numeric: tuple[str, ...]) -> list[dict]:
        out = []
        for r in rows:
            rr = dict(r)
            for c in numeric:
                if c in rr and isinstance(rr[c], (int, float)):
                    rr[c] = round(float(rr[c]), 4)
            out.append(rr)
        return out

    beams_rows = _round_rows(
        sorted(beams_by_section.values(), key=lambda r: r["length"], reverse=True), ("length", "mass")
    )
    plates_rows = _round_rows(
        sorted(plates_by_thk.values(), key=lambda r: r["area"], reverse=True), ("area", "mass", "thickness")
    )
    pipe_rows = _round_rows(sorted(pipe_by_size.values(), key=lambda r: r["length"], reverse=True), ("length", "mass"))
    duct_rows = _round_rows(
        sorted(duct_by_size.values(), key=lambda r: r["length"], reverse=True), ("length", "area", "mass")
    )
    tray_rows = _round_rows(sorted(tray_by_size.values(), key=lambda r: r["length"], reverse=True), ("length", "mass"))

    total = _Acc()
    for k in DISCIPLINE_ORDER:
        total.mass += disc_acc[k].mass
        total.moment += disc_acc[k].moment
        total.count += disc_acc[k].count

    disciplines = []
    for k in DISCIPLINE_ORDER:
        acc = disc_acc[k]
        disciplines.append(
            {
                "key": k,
                "name": DISCIPLINE_NAMES[k],
                "mass": acc.mass_t(),
                "cog": acc.cog(),
                "count": acc.count,
            }
        )

    if not np.isfinite(bbox_min).all():
        bbox = [0.0, 0.0, 0.0]
    else:
        bbox = [round(float(v), 3) for v in (bbox_max - bbox_min)]

    major.sort(key=lambda t: t[2], reverse=True)
    major_items = [
        {"name": name, "discipline": disc, "mass": round(mass, 4), "cog": cog} for name, disc, mass, cog in major[:12]
    ]

    joints = _joints_takeoff(part)

    return {
        "schema_version": SCHEMA_VERSION,
        "source_name": source_name,
        "units": {"length": "m", "mass": "tonne", "area": "m2"},
        "objects": n_objects,
        "total_mass": total.mass_t(),
        "total_cog": total.cog(),
        "bbox": bbox,
        "disciplines": disciplines,
        "structural": {
            "mass": disc_acc["structural"].mass_t(),
            "beams": beams_rows,
            "plates": plates_rows,
        },
        "piping": {
            "mass": disc_acc["piping"].mass_t(),
            "segments": pipe_rows,
            "fittings": [{"name": "Elbows", "count": n_elbows}, {"name": "Flanges", "count": n_flanges}],
        },
        "hvac": {
            "mass": disc_acc["hvac"].mass_t(),
            "segments": duct_rows,
            "fittings": [],
        },
        "electrical": {
            "mass": disc_acc["electrical"].mass_t(),
            "trays": tray_rows,
            "cables": [],
        },
        "joints": joints,
        "major_items": major_items,
    }


# --------------------------------------------------------------------------- #
# Exports: whole-model xlsx workbook + per-tab CSV                             #
# --------------------------------------------------------------------------- #
# The take-off tables are heterogeneous (a different column set per discipline,
# several tables on one discipline sheet), so the workbook is authored directly
# with openpyxl — the same library ``ada.serialize.xlsx`` builds on — rather than
# through the model-class-keyed WorkbookSerializer (one sheet per registered
# pydantic type), which does not fit multi-table sheets. The result matches the
# viewer's six tabs one-for-one: Overview, COGs, Structural, Piping, HVAC,
# Electrical.

# (label, stats-path, columns) table specs. ``path`` walks the stats dict.
_TAB_TABLES: dict[str, list[tuple[str, tuple[str, ...], list[tuple[str, str]]]]] = {
    "overview": [
        ("Mass by discipline", ("disciplines",), [("name", "Discipline"), ("mass", "Mass (t)"), ("count", "Objects")]),
    ],
    "cogs": [
        (
            "COG by discipline",
            ("disciplines",),
            [("name", "Discipline"), ("mass", "Mass (t)"), ("cog.0", "X (m)"), ("cog.1", "Y (m)"), ("cog.2", "Z (m)")],
        ),
        (
            "Major items",
            ("major_items",),
            [
                ("name", "Item"),
                ("discipline", "Discipline"),
                ("mass", "Mass (t)"),
                ("cog.0", "X (m)"),
                ("cog.1", "Y (m)"),
                ("cog.2", "Z (m)"),
            ],
        ),
    ],
    "structural": [
        (
            "Beams by section",
            ("structural", "beams"),
            [("section", "Section"), ("count", "Count"), ("length", "Length (m)"), ("mass", "Mass (t)")],
        ),
        (
            "Plates by thickness",
            ("structural", "plates"),
            [("label", "Plate"), ("count", "Count"), ("area", "Area (m²)"), ("mass", "Mass (t)")],
        ),
    ],
    "piping": [
        (
            "Segments by size",
            ("piping", "segments"),
            [("size", "Size"), ("segments", "Segments"), ("length", "Length (m)"), ("mass", "Mass (t)")],
        ),
        ("Fittings", ("piping", "fittings"), [("name", "Fitting"), ("count", "Count")]),
    ],
    "hvac": [
        (
            "Duct segments by size",
            ("hvac", "segments"),
            [
                ("size", "Size"),
                ("segments", "Segments"),
                ("length", "Length (m)"),
                ("area", "Area (m²)"),
                ("mass", "Mass (t)"),
            ],
        ),
        ("Fittings", ("hvac", "fittings"), [("name", "Fitting"), ("count", "Count")]),
    ],
    "electrical": [
        (
            "Cable tray by width",
            ("electrical", "trays"),
            [("size", "Width"), ("segments", "Segments"), ("length", "Length (m)"), ("mass", "Mass (t)")],
        ),
        ("Cables by type", ("electrical", "cables"), [("type", "Type"), ("length", "Length (m)")]),
    ],
    "joints": [
        (
            "Joints by type",
            ("joints", "by_type"),
            [("name", "Type"), ("count", "Count")],
        ),
        (
            "Joints",
            ("joints", "items"),
            [
                ("name", "Name"),
                ("type", "Type"),
                ("plates", "Plates"),
                ("welds", "Welds"),
                ("centre.0", "X (m)"),
                ("centre.1", "Y (m)"),
                ("centre.2", "Z (m)"),
            ],
        ),
    ],
}

_TAB_SHEET_NAMES: dict[str, str] = {
    "overview": "Overview",
    "cogs": "COGs",
    "structural": "Structural",
    "piping": "Piping",
    "hvac": "HVAC",
    "electrical": "Electrical",
    "joints": "Joints",
}
_EXPORT_TABS: tuple[str, ...] = ("overview", "cogs", "structural", "piping", "hvac", "electrical", "joints")


def _dig(row: dict, path: str):
    """Resolve a dotted column key against a row (``"cog.0"`` -> ``row['cog'][0]``)."""
    cur = row
    for part in path.split("."):
        if isinstance(cur, (list, tuple)):
            cur = cur[int(part)]
        elif isinstance(cur, dict):
            cur = cur.get(part)
        else:
            return None
    return cur


def _rows_at(stats: dict, path: tuple[str, ...]) -> list[dict]:
    cur = stats
    for p in path:
        cur = cur.get(p, {}) if isinstance(cur, dict) else {}
    return cur if isinstance(cur, list) else []


def takeoff_to_xlsx_bytes(stats: dict) -> bytes:
    """Serialize a take-off ``stats`` dict to a whole-model Excel workbook (bytes):
    one sheet per discipline plus COGs and Overview, matching the viewer tabs."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)

    for tab in _EXPORT_TABS:
        ws = wb.create_sheet(_TAB_SHEET_NAMES[tab])
        r = 1
        if tab == "overview":
            ws.cell(r, 1, "Metric").font = bold
            ws.cell(r, 2, "Value").font = bold
            r += 1
            for label, val in (
                ("Total mass (t)", stats.get("total_mass")),
                ("Objects", stats.get("objects")),
                ("Disciplines", len([d for d in stats.get("disciplines", []) if d.get("count")])),
                ("Bounding box (m)", " × ".join(str(v) for v in stats.get("bbox", []))),
                ("COG X (m)", (stats.get("total_cog") or [None])[0]),
                ("COG Y (m)", (stats.get("total_cog") or [None, None])[1]),
                ("COG Z (m)", (stats.get("total_cog") or [None, None, None])[2]),
            ):
                ws.cell(r, 1, label)
                ws.cell(r, 2, val)
                r += 1
            r += 1
        for title, path, cols in _TAB_TABLES[tab]:
            ws.cell(r, 1, title).font = bold
            r += 1
            for ci, (_key, header) in enumerate(cols, start=1):
                ws.cell(r, ci, header).font = bold
            r += 1
            for row in _rows_at(stats, path):
                for ci, (key, _header) in enumerate(cols, start=1):
                    ws.cell(r, ci, _dig(row, key))
                r += 1
            r += 1  # blank spacer between tables

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def takeoff_to_csv(stats: dict, tab: str) -> str:
    """Serialize the tables of one viewer tab to CSV text (the active-tab export)."""
    import csv
    import io

    tab = tab if tab in _TAB_TABLES else "overview"
    buf = io.StringIO()
    writer = csv.writer(buf)
    for title, path, cols in _TAB_TABLES[tab]:
        writer.writerow([title])
        writer.writerow([header for _key, header in cols])
        for row in _rows_at(stats, path):
            writer.writerow([_dig(row, key) for key, _header in cols])
        writer.writerow([])
    return buf.getvalue()
