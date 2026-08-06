"""Generic openpyxl styling helpers — header look, dropdown derivation, column
widths. No domain knowledge; safe to reuse across models."""

from __future__ import annotations

import enum
from typing import TYPE_CHECKING, Any, Literal, get_args, get_origin

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["header_style", "dropdown_values", "tune_column_width", "DROPDOWN_LIMIT"]

# openpyxl caps an inline data-validation list formula at 255 characters.
DROPDOWN_LIMIT = 255

# Header fill/font are built lazily on first use so that merely importing this
# module (e.g. via ``ada.topo_model``) does not require openpyxl — only actually
# writing a workbook does. Cached after the first call.
_HEADER_STYLE: tuple | None = None


def header_style(cell) -> None:
    global _HEADER_STYLE
    if _HEADER_STYLE is None:
        from openpyxl.styles import Font, PatternFill

        _HEADER_STYLE = (
            PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            Font(bold=True),
        )
    cell.fill, cell.font = _HEADER_STYLE


def dropdown_values(core: Any) -> list[str]:
    """The allowed values for a ``Literal[...]`` (multi-value) or ``Enum`` field,
    as strings — used to build a cell dropdown. Empty for anything else."""
    if get_origin(core) is Literal:
        args = [a for a in get_args(core) if a is not None]
        return [str(a) for a in args] if len(args) > 1 else []
    if isinstance(core, type) and issubclass(core, enum.Enum):
        return [m.name for m in core]  # by member name, matching EnumCodec
    return []


def tune_column_width(ws: Worksheet, columns: list[str]) -> None:
    from openpyxl.utils import get_column_letter

    for c, name in enumerate(columns, start=1):
        # Upper-case letters read wider; a light heuristic keeps headers legible.
        width = min(48, max(10, int(len(name) * 1.1) + 2))
        ws.column_dimensions[get_column_letter(c)].width = width
