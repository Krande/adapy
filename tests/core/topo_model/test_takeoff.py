"""Quantity take-off (:mod:`ada.topo_model.takeoff`) — the data behind the
viewer's Stats panel.

Numbers are pinned against the canonical IPE200/HEB200/HP140x8 steel demo
(``build_topo_model``), whose member counts are themselves pinned in
``test_steel_stru``. The systems demo exercises the piping + electrical
branches; the export helpers assert a valid multi-sheet workbook + CSV.
"""

from __future__ import annotations

import io

import pytest

import ada
from ada.topo_model import build_topo_model
from ada.topo_model.build import build_topo_model_with_systems
from ada.topo_model.takeoff import (
    classify_discipline,
    model_takeoff,
    takeoff_to_csv,
    takeoff_to_xlsx_bytes,
)


@pytest.fixture(scope="module")
def steel_takeoff() -> dict:
    return model_takeoff(build_topo_model(), source_name="TopoModelDemo")


def test_steel_headline_numbers(steel_takeoff):
    t = steel_takeoff
    assert t["schema_version"] == 2
    assert t["source_name"] == "TopoModelDemo"
    assert t["units"] == {"length": "m", "mass": "tonne", "area": "m2"}
    # 72 physical objects: 68 beams (48 HP + 14 IPE + 6 HEB) + 4 deck plates.
    assert t["objects"] == 72
    assert t["total_mass"] == pytest.approx(12.7763, abs=1e-3)
    assert t["total_cog"] == pytest.approx([5.0, 2.5185, 1.4722], abs=1e-3)
    assert t["bbox"] == pytest.approx([10.0, 5.0, 3.0], abs=1e-3)


def test_structural_only_model_has_empty_joints(steel_takeoff):
    # A model compiled without a detailing engine carries no connection joints.
    j = steel_takeoff["joints"]
    assert j == {"count": 0, "by_type": [], "items": []}


def test_detailed_model_joints_takeoff():
    # With the built-in detailing engine applied, the take-off surfaces the per-
    # type roll-up (the Detailing tab's "N detected") and a per-instance table.
    from ada.topo_model.detailing import detail

    a = build_topo_model()
    detail(a, {})
    t = model_takeoff(a, source_name="Detailed")
    j = t["joints"]
    # 6 column base plates only: end plates dropped, HP stringers excluded, and
    # every girder node carries a column (no bare girder gusset) — see test_detailing.
    assert j["count"] == 6
    by_slug = {r["slug"]: r["count"] for r in j["by_type"]}
    assert by_slug == {"column_base_plate": 6}
    assert all(r.get("name") for r in j["by_type"])
    # Every instance carries framed members, plate/weld counts and a node centre.
    bp = next(it for it in j["items"] if it["slug"] == "column_base_plate")
    assert bp["plates"] == 1 and bp["welds"] >= 1
    assert bp["members"]
    assert bp["centre"] is not None and len(bp["centre"]) == 3


def test_steel_disciplines_are_structural_only(steel_takeoff):
    by_key = {d["key"]: d for d in steel_takeoff["disciplines"]}
    assert set(by_key) == {"structural", "piping", "hvac", "electrical"}
    assert by_key["structural"]["mass"] == pytest.approx(12.7763, abs=1e-3)
    assert by_key["structural"]["count"] == 72
    for empty in ("piping", "hvac", "electrical"):
        assert by_key[empty]["mass"] == 0.0
        assert by_key[empty]["count"] == 0


def test_steel_beams_by_section(steel_takeoff):
    rows = {r["section"]: r for r in steel_takeoff["structural"]["beams"]}
    assert rows["HP140x8"]["count"] == 48
    assert rows["HP140x8"]["length"] == pytest.approx(240.0, abs=1e-6)
    assert rows["HP140x8"]["mass"] == pytest.approx(2.365, abs=1e-3)
    assert rows["IPE200"]["count"] == 14
    assert rows["IPE200"]["length"] == pytest.approx(70.0, abs=1e-6)
    assert rows["IPE200"]["mass"] == pytest.approx(1.4973, abs=1e-3)
    assert rows["HEB200"]["count"] == 6
    assert rows["HEB200"]["length"] == pytest.approx(18.0, abs=1e-6)
    assert rows["HEB200"]["mass"] == pytest.approx(1.064, abs=1e-3)
    # Sorted largest-length first.
    assert [r["section"] for r in steel_takeoff["structural"]["beams"]] == ["HP140x8", "IPE200", "HEB200"]


def test_steel_plates_by_thickness(steel_takeoff):
    plates = steel_takeoff["structural"]["plates"]
    assert len(plates) == 1
    pl = plates[0]
    assert pl["label"] == "PL10"
    assert pl["thickness"] == pytest.approx(0.01, abs=1e-9)
    assert pl["count"] == 4
    assert pl["area"] == pytest.approx(100.0, abs=1e-6)  # 4 x 5x5 decks
    assert pl["mass"] == pytest.approx(7.85, abs=1e-3)  # 100 m2 x 0.01 m x 7850 kg/m3


def test_major_items_sorted_desc(steel_takeoff):
    masses = [m["mass"] for m in steel_takeoff["major_items"]]
    assert masses == sorted(masses, reverse=True)
    assert all(m["discipline"] == "structural" for m in steel_takeoff["major_items"])


def test_classifier():
    from ada.api.piping.base_piping import PipeSegStraight

    bm = ada.Beam("b", (0, 0, 0), (1, 0, 0), "IPE200")
    assert classify_discipline(bm) == "structural"

    class _Tagged:
        metadata = {"segment_ifc_class": "IfcDuctSegment"}

    assert classify_discipline(_Tagged()) == "hvac"
    _Tagged.metadata = {"segment_ifc_class": "IfcCableSegment"}
    assert classify_discipline(_Tagged()) == "electrical"
    _Tagged.metadata = {"segment_ifc_class": "IfcPipeSegment"}
    assert classify_discipline(_Tagged()) == "piping"

    sec = ada.Section("p", "PIPE", r=0.05, wt=0.005)
    seg = PipeSegStraight("s", (0, 0, 0), (1, 0, 0), sec, ada.Material("S355"))
    assert classify_discipline(seg) == "piping"

    assert classify_discipline(object()) == "other"


def test_systems_demo_populates_piping_and_electrical():
    t = model_takeoff(build_topo_model_with_systems())
    by_key = {d["key"]: d for d in t["disciplines"]}
    assert by_key["piping"]["mass"] > 0.0
    assert by_key["piping"]["count"] > 0
    assert by_key["electrical"]["mass"] > 0.0
    assert by_key["electrical"]["count"] > 0
    # Piping take-off table + fittings roll-up are present.
    assert t["piping"]["segments"], "expected pipe segments grouped by size"
    assert sum(r["length"] for r in t["piping"]["segments"]) > 0.0
    fittings = {f["name"]: f["count"] for f in t["piping"]["fittings"]}
    assert fittings["Elbows"] > 0
    # Electrical trays grouped by width.
    assert t["electrical"]["trays"], "expected cable-tray segments grouped by width"


def test_xlsx_export_is_valid_workbook(steel_takeoff):
    from openpyxl import load_workbook

    data = takeoff_to_xlsx_bytes(steel_takeoff)
    assert data[:2] == b"PK"  # a zip container
    assert len(data) > 0
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Overview", "COGs", "Structural", "Piping", "HVAC", "Electrical", "Joints"]
    # The Structural sheet carries the beams-by-section rows.
    text = "\n".join(
        str(c.value) for row in wb["Structural"].iter_rows() for c in row if c.value is not None
    )
    assert "HP140x8" in text and "IPE200" in text and "HEB200" in text


def test_csv_export_active_tab(steel_takeoff):
    csv = takeoff_to_csv(steel_takeoff, "structural")
    assert "Beams by section" in csv
    assert "Section,Count,Length (m),Mass (t)" in csv
    assert "HP140x8" in csv
    assert "Plates by thickness" in csv
    # Unknown tab falls back to overview rather than erroring.
    assert "Discipline" in takeoff_to_csv(steel_takeoff, "nope")
