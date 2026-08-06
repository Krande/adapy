"""WorkbookSerializer: register pydantic models, then write/read a workbook.

The engine is generic — it holds no domain model list and no ``isinstance``
ladder. A caller registers each model (its :class:`SheetSpec` is derived from the
class ClassVars) and a codec registry does the type<->cell work. ``read`` returns
a ``{model_cls: [instances]}`` map; assembling those into a domain document is the
caller's job, keeping the serializer reusable across repos.

Extension points:
* **codecs** — pass/register a :class:`CodecRegistry` to teach new types.
* **per-model hooks** (optional methods on the model, all duck-typed):
  ``excel_should_skip_row(cls, raw: dict) -> bool`` — drop a row on read;
  ``excel_pre_write(self) -> dict`` — override/derive cell values on write;
  ``excel_post_read(cls, values: dict) -> dict`` — adjust parsed values before
  the model is constructed.
* **layouts** — register a handler for a new ``ORIENTATION`` via
  :meth:`register_layout`.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any, Callable, Iterable

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

from .codecs import CodecRegistry, core_type, default_registry
from .spec import SheetSpec, field_description
from .styling import DROPDOWN_LIMIT, dropdown_values, header_style, tune_column_width

__all__ = ["WorkbookSerializer", "SCHEMA_VERSION"]

SCHEMA_VERSION = "1"
_STAMP = f"adapy.serialize.xlsx/{SCHEMA_VERSION}"


class WorkbookSerializer:
    def __init__(self, codecs: CodecRegistry | None = None) -> None:
        self.codecs = codecs or default_registry()
        self._specs: dict[type, SheetSpec] = {}
        self._layouts: dict[str, "_Layout"] = {
            "HORIZONTAL": _Horizontal(),
            "VERTICAL": _Vertical(),
        }

    # -- registration ------------------------------------------------------
    def register(self, model_cls: type, spec: SheetSpec | None = None) -> "WorkbookSerializer":
        self._specs[model_cls] = spec or SheetSpec.from_model(model_cls)
        return self

    def register_layout(self, orientation: str, layout: "_Layout") -> "WorkbookSerializer":
        self._layouts[orientation.upper()] = layout
        return self

    def _layout_for(self, spec: SheetSpec) -> "_Layout":
        try:
            return self._layouts[spec.orientation.upper()]
        except KeyError:
            raise ValueError(
                f"{spec.model.__name__}: no layout registered for ORIENTATION {spec.orientation!r} "
                f"(known: {sorted(self._layouts)})"
            ) from None

    # -- write -------------------------------------------------------------
    def write(self, instances: Iterable[Any], path: str | Path) -> None:
        """Write every instance to its model's sheet. Instances of an
        unregistered type raise (never silently skipped)."""
        from openpyxl import Workbook

        grouped: dict[type, list[Any]] = {}
        for obj in instances:
            if type(obj) not in self._specs:
                raise ValueError(f"{type(obj).__name__} is not registered with this serializer")
            grouped.setdefault(type(obj), []).append(obj)

        wb = Workbook()
        wb.remove(wb.active)  # drop the default empty sheet
        # Emit a sheet for every registered model (empty ones too), so the file is
        # a complete template even when a model has no rows yet.
        for model_cls, spec in self._specs.items():
            ws = wb.create_sheet(title=spec.sheet_name)
            self._layout_for(spec).write(ws, spec, grouped.get(model_cls, []), self.codecs)
        wb.properties.creator = _STAMP
        wb.save(path)

    # -- read --------------------------------------------------------------
    def read(self, path: str | Path) -> dict[type, list[Any]]:
        """Parse a workbook into ``{model_cls: [instances]}`` for every
        registered model. A registered model whose sheet is absent yields ``[]``."""
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        out: dict[type, list[Any]] = {}
        for model_cls, spec in self._specs.items():
            if spec.sheet_name not in wb.sheetnames:
                out[model_cls] = []
                continue
            out[model_cls] = self._layout_for(spec).read(wb[spec.sheet_name], spec, self.codecs)
        return out


# ---------------------------------------------------------------------------
# Layout handlers
# ---------------------------------------------------------------------------
def _cell_value(codecs: CodecRegistry, model_cls: type, name: str, obj: Any) -> Any:
    codec = codecs.resolve(model_cls.model_fields[name], name)
    return codec.to_cell(getattr(obj, name))


def _row_values(codecs: CodecRegistry, spec: SheetSpec, obj: Any) -> dict[str, Any]:
    override = obj.excel_pre_write() if hasattr(obj, "excel_pre_write") else {}
    values: dict[str, Any] = {}
    for name in spec.columns:
        if name in override:
            values[name] = override[name]
        else:
            values[name] = _cell_value(codecs, spec.model, name, obj)
    return values


def _parse_row(codecs: CodecRegistry, spec: SheetSpec, raw: dict[str, Any]) -> Any | None:
    model_cls = spec.model
    if hasattr(model_cls, "excel_should_skip_row") and model_cls.excel_should_skip_row(raw):
        return None
    values: dict[str, Any] = {}
    for name, cell in raw.items():
        field_info = model_cls.model_fields.get(name)
        if field_info is None:
            continue  # unknown column — ignore
        codec = codecs.resolve(field_info, name)
        parsed = codec.from_cell(cell, core_type(field_info.annotation))
        if parsed is not None:  # None -> let the field default apply
            values[name] = parsed
    if hasattr(model_cls, "excel_post_read"):
        values = model_cls.excel_post_read(values)
    return model_cls(**values)


class _Layout:
    def write(self, ws: Worksheet, spec: SheetSpec, items: list[Any], codecs: CodecRegistry) -> None:
        raise NotImplementedError

    def read(self, ws: Worksheet, spec: SheetSpec, codecs: CodecRegistry) -> list[Any]:
        raise NotImplementedError


class _Horizontal(_Layout):
    """Row 1 = headers, one instance per subsequent row."""

    def write(self, ws, spec, items, codecs):
        from openpyxl.comments import Comment
        from openpyxl.utils import get_column_letter

        cols = spec.columns
        hidden = set(spec.hidden)
        for c, name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c, value=name)
            header_style(cell)
            desc = field_description(spec.model, name)
            if desc:
                cell.comment = Comment(str(desc), _STAMP)
            _add_dropdown(ws, spec, name, col_letter=get_column_letter(c), first_row=2)
            if name in hidden:
                ws.column_dimensions[get_column_letter(c)].hidden = True
        for r, obj in enumerate(items, start=2):
            values = _row_values(codecs, spec, obj)
            for c, name in enumerate(cols, start=1):
                ws.cell(row=r, column=c, value=values[name])
        if spec.tab_color:
            ws.sheet_properties.tabColor = spec.tab_color
        tune_column_width(ws, cols)

    def read(self, ws, spec, codecs):
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return []
        names = [str(h).strip() if h is not None else None for h in header]
        out: list[Any] = []
        for row in rows:
            if all(_blank(v) for v in row):
                continue
            raw = {names[i]: row[i] for i in range(len(names)) if names[i]}
            parsed = _parse_row(codecs, spec, raw)
            if parsed is not None:
                out.append(parsed)
        return out


class _Vertical(_Layout):
    """Column A = field name, column B = value; the whole sheet is ONE instance."""

    def write(self, ws, spec, items, codecs):
        from openpyxl.comments import Comment

        if len(items) > 1:
            raise ValueError(f"{spec.sheet_name}: VERTICAL layout holds a single instance, got {len(items)}")
        cols = spec.columns
        hidden = set(spec.hidden)
        values = _row_values(codecs, spec, items[0]) if items else {}
        for r, name in enumerate(cols, start=1):
            key = ws.cell(row=r, column=1, value=name)
            header_style(key)
            desc = field_description(spec.model, name)
            if desc:
                key.comment = Comment(str(desc), _STAMP)
            ws.cell(row=r, column=2, value=values.get(name))
            if name in hidden:
                ws.row_dimensions[r].hidden = True
        if spec.tab_color:
            ws.sheet_properties.tabColor = spec.tab_color

    def read(self, ws, spec, codecs):
        raw: dict[str, Any] = {}
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            name = str(row[0]).strip()
            raw[name] = row[1] if len(row) > 1 else None
        if all(_blank(v) for v in raw.values()):
            return []
        parsed = _parse_row(codecs, spec, raw)
        return [parsed] if parsed is not None else []


def _add_dropdown(ws: Worksheet, spec: SheetSpec, name: str, col_letter: str, first_row: int) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation

    field_info = spec.model.model_fields.get(name)
    if field_info is None:
        return
    values = dropdown_values(core_type(field_info.annotation))
    if not values:
        return
    formula = '"' + ",".join(values) + '"'
    if len(formula) > DROPDOWN_LIMIT:
        return  # openpyxl inline-list limit; skip rather than corrupt the sheet
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(f"{col_letter}{first_row}:{col_letter}1048576")
    ws.add_data_validation(dv)


def _blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


_LayoutFactory = Callable[[], _Layout]
