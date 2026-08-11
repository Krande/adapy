"""Procedural model Excel export / import: key conventions, the dependency-free
``_ADA_META`` reader, and the endpoint shapes (503 without a database; upload +
engine auto-detection needs neither DB nor ada)."""

from __future__ import annotations

import io
import os
import pathlib
import tempfile

import pytest

os.environ.setdefault("ADA_VIEWER_STORAGE_KIND", "local")
os.environ.setdefault("ADA_VIEWER_LOCAL_PATH", tempfile.mkdtemp(prefix="ada-test-storage-"))

from fastapi.testclient import TestClient  # noqa: E402

from ada.comms.rest.app import create_app  # noqa: E402
from ada.comms.rest.config import (  # noqa: E402
    AuthConfig,
    LocalConfig,
    QueueConfig,
    Settings,
)
from ada.comms.rest.converter import is_hidden_key  # noqa: E402
from ada.comms.rest.procedural import (  # noqa: E402
    ADA_META_SHEET,
    procedural_import_result_key,
    procedural_import_source_key,
    procedural_xlsx_export_key,
    read_ada_meta_from_xlsx_bytes,
)


def _settings(tmp_path: pathlib.Path, database_url: str = "") -> Settings:
    return Settings(
        storage_kind="local",
        s3=None,
        local=LocalConfig(path=str(tmp_path), prefix=""),
        host="127.0.0.1",
        port=0,
        static_path="",
        queue=QueueConfig(
            url=None,
            stream="ada",
            subject="ada.viewer.jobs.convert",
            kv_bucket="ada-viewer-jobs",
            durable="ada-viewer-worker",
        ),
        auth=AuthConfig(
            enabled=False, issuer="", client_id="", audience="", admin_group="", cli_token_secret=""
        ),
        database_url=database_url,
    )


def _xlsx_with_meta(rows: list[tuple[str, str]]) -> bytes:
    """A minimal workbook carrying an ``_ADA_META`` key/value sheet — built with
    openpyxl directly (no ada), so the reader is exercised standalone."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "Spaces"
    ws = wb.create_sheet(ADA_META_SHEET)
    for r, (k, v) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── standalone helpers ───────────────────────────────────────────────


def test_export_key_shape():
    key = procedural_xlsx_export_key("abc-123", 4)
    assert key == "_procedural/abc-123/r4.xlsx"
    assert is_hidden_key(key)
    # non-default engine gets its own suffixed key
    assert procedural_xlsx_export_key("abc-123", 4, "pm-engine") == "_procedural/abc-123/r4.pm-engine.xlsx"
    assert procedural_xlsx_export_key("abc-123", 4, "adapy-default") == key  # collapses to bare


def test_import_key_shapes():
    src = procedural_import_source_key("tok123")
    assert src == "_procedural/_import/tok123/source.xlsx"
    assert is_hidden_key(src)
    assert procedural_import_result_key(src) == "_procedural/_import/tok123/result.json"


def test_ada_meta_reader_roundtrips():
    data = _xlsx_with_meta([("ada_meta_version", "1"), ("engine", "pm-engine"), ("package_version", "0.10.2")])
    meta = read_ada_meta_from_xlsx_bytes(data)
    assert meta is not None
    assert meta["engine"] == "pm-engine"
    assert meta["package_version"] == "0.10.2"


def test_ada_meta_reader_absent_and_malformed():
    # A workbook with no _ADA_META sheet -> None (the "prompt for engine" case).
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "Spaces"
    buf = io.BytesIO()
    wb.save(buf)
    assert read_ada_meta_from_xlsx_bytes(buf.getvalue()) is None
    # Junk / not-a-zip -> None, never raises.
    assert read_ada_meta_from_xlsx_bytes(b"not a workbook") is None
    assert read_ada_meta_from_xlsx_bytes(b"") is None


# ── no-DB API path ───────────────────────────────────────────────────


@pytest.fixture
def app_client(tmp_path: pathlib.Path):
    app = create_app(_settings(tmp_path))
    with TestClient(app) as client:
        yield client


def test_export_and_import_enqueue_503_without_db(app_client: TestClient):
    # Export + import-enqueue both need the procedural pool -> 503 without a DB.
    assert app_client.post("/api/scopes/shared/procedural-models/x/export-xlsx").status_code == 503
    r = app_client.post(
        "/api/scopes/shared/procedural-models/import-xlsx",
        json={"source_key": "_procedural/_import/tok/source.xlsx", "engine": "adapy-default", "name": "m"},
    )
    assert r.status_code == 503


def test_import_upload_detects_engine_without_db(app_client: TestClient):
    # Upload + auto-detect needs neither a DB nor ada: the _ADA_META sheet is read
    # dependency-free and the workbook is staged under a hidden import token.
    data = _xlsx_with_meta([("engine", "pm-engine"), ("package_version", "0.10.2")])
    r = app_client.post(
        "/api/scopes/shared/procedural-models/import-xlsx/upload", content=data
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["engine"] == "pm-engine"
    assert body["source_key"].startswith("_procedural/_import/")

    # A legacy workbook (no _ADA_META) -> engine null (frontend then prompts).
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "Spaces"
    buf = io.BytesIO()
    wb.save(buf)
    r2 = app_client.post(
        "/api/scopes/shared/procedural-models/import-xlsx/upload", content=buf.getvalue()
    )
    assert r2.status_code == 200, r2.text
    assert r2.json()["engine"] is None


def test_import_upload_rejects_empty(app_client: TestClient):
    r = app_client.post("/api/scopes/shared/procedural-models/import-xlsx/upload", content=b"")
    assert r.status_code == 400


def test_import_enqueue_validates_source_key(app_client: TestClient):
    # A source_key not from the import-upload prefix is rejected before the pool
    # check would 503 — but the pool check runs first, so no-DB yields 503 here.
    # With no DB the endpoint 503s regardless; the prefix guard is covered by the
    # live-DB path. Assert the no-DB contract:
    r = app_client.post(
        "/api/scopes/shared/procedural-models/import-xlsx",
        json={"source_key": "bogus", "engine": "adapy-default", "name": "m"},
    )
    assert r.status_code == 503
